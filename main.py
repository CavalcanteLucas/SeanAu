from argparse import ArgumentParser
from xlsxwriter import Workbook
from requests import get
from openpyxl import load_workbook
from distutils import util


CLICKIDS_FILE = 'clickid.xlsx'
OUTPUT_FILENAME = 'output.xlsx'
URL = 'http://firequantummedia.com/arm.php'
API_KEY = '1000001eae3f94064deb69a4f90629d651e69ac'


def build_params(clickid):
    return {
        'api_key': API_KEY,
        'action': 'clickinfo@get',
        'clickid': clickid,
    }


def make_request(clickid, verbose_value):
    if verbose_value:
        print(f' [x] Making request for clickId: {clickid} ...')
    response = get(url=URL, params=build_params(clickid))
    return response.json()


def get_clickids():
    wb = load_workbook(filename=CLICKIDS_FILE)
    ws = wb['Sheet1']
    clickids = []
    for line in ws.iter_cols(min_row=2, max_col=1):
        for cell in line:
            clickd = cell.value
            if clickd:
                clickids.append(cell.value)
    return clickids


def get_data(clickids, verbose_value):
    header = []
    values = []
    for id in clickids:
        data = make_request(clickid=id, verbose_value=verbose_value)
        for item in data.items():
            if item[0] == 'click':
                data_from_json = item[1]
                if header:
                    assert header == [*data_from_json]
                else:
                    header = [*data_from_json]
                values.append([*data_from_json.values()])
    return [header, values]


def write_into_xlsx(rows, output_file):
    workbook = Workbook(output_file)
    worksheet = workbook.add_worksheet()
    row_index = column_index = 0
    for row in rows:
        column_index = 0
        # header
        if row_index == 0:
            for item in row:
                worksheet.write(row_index, column_index, item)
                column_index += 1
        # values
        else:
            for item in row:
                for column in item:
                    worksheet.write(row_index, column_index, column)
                    column_index += 1
                column_index = 0
                row_index += 1
        row_index += 1
    workbook.close()


def mainscript(verbose_value):
    clickids = get_clickids()
    rows = get_data(clickids, verbose_value)
    write_into_xlsx(rows=rows, output_file=OUTPUT_FILENAME)


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument(
        '--verbose',
        type=util.strtobool,
        default='True',
        help='Verbose option. Default = True',
    )
    args = parser.parse_args()

    verbose_value = args.verbose

    mainscript(verbose_value)
